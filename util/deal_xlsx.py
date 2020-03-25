# !/usr/bin/env python
# -*- coding=utf-8 -*-

import openpyxl


class DealXlsx(object):

    file_name = ""
    header = []
    body = []
    # object_yaml = None
    wb = None
    sh = None


    # def new_xlsx_file(self):
    #     f = open(self.file_name, 'w', encoding='utf-8')
    #     return f

    def new_work_book(self):
        self.wb = openpyxl.Workbook()

    def new_sheet(self, sheet_name):
        self.wb.create_sheet(title=sheet_name)

    def save_xlsx_file(self, file_name):
        self.wb.save(filename=file_name)

    def open_work_book(self, file_name):
        self.wb = openpyxl.load_workbook(file_name)

    def get_sheet(self, sheet_name):
        return self.wb[sheet_name]

    # def csv_writer(self, f):
    #     return csv.writer(f)
    #
    # def add_body(self, row):
    #     self.body.append(row)

    def sheet_append_row(self, sheet, row):
        sheet.append(row)

    # def close_file(self, f):
    #     f.close()

    def object2xlsx(self):
        self.new_work_book()
        self.sh = self.wb.active

        if self.header and self.body:
            self.sheet_append_row(self.sh, self.header)
            for row in self.body:
                for i in range(len(row)):
                    if isinstance(row[i], str):
                        row[i] = row[i].replace("\n", "_x000D_")
                self.sheet_append_row(self.sh, row)

        if self.file_name:
            self.save_xlsx_file(self.file_name)